'''
Author: zolty zolty@qq.com
Date: 2025-04-27 09:56:43
LastEditors: zolty zolty@qq.com
LastEditTime: 2025-04-29 14:25:53
FilePath: /ads-player/tools/temp_excel.py
Description: 这是默认设置,请设置`customMade`, 打开koroFileHeader查看配置 进行设置: https://github.com/OBKoro1/koro1FileHeader/wiki/%E9%85%8D%E7%BD%AE
'''
from openpyxl import load_workbook
import pandas as pd

# 定义车型列表
vmodel = ["小鹏P7+","问界M7","智己LS6","理想L6","埃安Y"]


# ============= 你需要填的内容 广州 =============
# test_entries = [
#     "广州_小鹏P7+_20250311_晚闲时_线路E",
#     "广州_小鹏P7+_20250313_早高峰_线路E",
#     "广州_问界M7_20250311_晚闲时_线路E",
#     "广州_问界M7_20250313_早高峰_线路E",
#     "广州_智己LS6_20250311_晚闲时_线路E",
#     "广州_智己LS6_20250313_早高峰_线路E",
#     "广州_理想L6_20250311_晚闲时_线路E",
#     "广州_理想L6_20250313_早高峰_线路E",
#     "广州_埃安Y_20250311_晚闲时_线路E",
#     "广州_埃安Y_20250313_早高峰_线路E"
# ]
# niuma  = ["朱煜","姜腾龙","胡鑫磊、甘雨、裴天箫"]

# ============= 你需要填的内容 深圳 =============
# test_entries = [
#     "深圳_小鹏P7+_20250317_晚闲时_线路E",
#     "深圳_小鹏P7+_20250319_早高峰_线路E",
#     "深圳_问界M7_20250317_晚闲时_线路E", 
#     "深圳_问界M7_20250319_早高峰_线路E",
#     "深圳_智己LS6_20250317_晚闲时_线路E",
#     "深圳_智己LS6_20250319_早高峰_线路E",
#     "深圳_理想L6_20250317_晚闲时_线路E", 
#     "深圳_理想L6_20250319_早高峰_线路E",
#     "深圳_埃安Y_20250317_晚闲时_线路E",
#     "深圳_埃安Y_20250319_早高峰_线路E"
# ]
# niuma  = ["朱煜","张子晗","朱煜、张子晗、裴天箫"]
# ============= 你需要填的内容 =============

# ============= 你需要填的内容 深圳 =============
test_entries = [
    "东莞_小鹏P7+_20250314_晚闲时_线路E",
    "东莞_小鹏P7+_20250316_早高峰_线路E",
    "东莞_问界M7_20250314_晚闲时_线路E",
    "东莞_问界M7_20250316_早高峰_线路E",
    "东莞_智己LS6_20250314_晚闲时_线路E",
    "东莞_智己LS6_20250316_早高峰_线路E",
    "东莞_理想L6_20250314_晚闲时_线路E",
    "东莞_理想L6_20250316_早高峰_线路E",
    "东莞_埃安Y_20250314_晚闲时_线路E",
    "东莞_埃安Y_20250316_早高峰_线路E"
]
niuma  = ["朱煜","张子晗","朱煜、张子晗、甘雨"]
# ============= 你需要填的内容 =============

          
# 加载模板文件
template_file = '汇总表_模板.xlsx'
wb_template = load_workbook(template_file)
ws_template = wb_template.active

prefix = '汇总表_'

# 遍历每个测试条目
for entry in test_entries:
    # 解析条目名称
    parts = entry.split('_')
    location, model, date, time_period, line = parts[0], parts[1], parts[2], parts[3], parts[4]
    
    # 创建一个新的工作簿副本
    wb_new = load_workbook(template_file)
    ws_new = wb_new.active
    
    # 设置 Sheet 名为测试条目名称的简短形式
    ws_new.title = prefix + entry  # 设置 Sheet 名
    
    # 填充基本信息到第一行
    ws_new['X3'] = entry  # 测试名称
    ws_new['X4'] = location  # 地点
    ws_new['X5'] = model  # 车型
    ws_new['X6'] = f"{date[:4]}/{date[4:6]}/{date[6:]}"  # 日期
    ws_new['X7'] = time_period  # 时段
    ws_new['X8'] = line  # 线路

    ws_new['X11'] = niuma[0] 
    ws_new['X12'] = niuma[1] 
    ws_new['X13'] = niuma[2]

    # 填充100行数据（假设从第10行开始是数据区域）
    for i in range(2, 100):
        ws_new[f'A{i}'] = i - 2   # 序号
        ws_new[f'F{i}'] = vmodel.index(model) + 1  # 车型序号
        ws_new[f'H{i}'] = location
        ws_new[f'I{i}'] = date
        ws_new[f'J{i}'] = line
        ws_new[f'K{i}'] = time_period
        ws_new[f'L{i}'] = "晴朗"

    for i in range(2, 10):
        ws_new[f'AA{i}'] = entry
        ws_new[f'AB{i}'] = model

    
    # 输出文件名
    output_filename = f"{prefix + entry}.xlsx"
    
    # 保存新文件
    wb_new.save(output_filename)

print("所有文件生成完毕")